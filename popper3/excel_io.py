from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


HEADERS = ["titulo", "descricao", "pontuacao", "veracidade"]
MAX_TITLE_CHARS = 32
MAX_GAME_DESCRIPTION_CHARS = 180


def create_template(path: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartas de jogo"
    ws.append(HEADERS)
    ws.append(["Exemplo de carta", "Descricao da carta com ate 180 caracteres", 10, "verdadeira"])
    ws.append(["Outra carta", "Descricao da carta com ate 180 caracteres", 5, "falsa"])
    for col, width in zip("ABCD", [34, 70, 14, 16], strict=True):
        ws.column_dimensions[col].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")
    wb.save(path)


def read_game_cards(path: str | Path) -> list[dict[str, object]]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    missing = [header for header in HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Colunas ausentes no Excel: {', '.join(missing)}")

    idx = {header: headers.index(header) for header in HEADERS}
    cards: list[dict[str, object]] = []
    errors: list[str] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        title = row[idx["titulo"]]
        if not title:
            continue
        title_text = str(title).strip()
        description_text = str(row[idx["descricao"]] or "").strip()
        if len(title_text) > MAX_TITLE_CHARS:
            errors.append(f"Linha {row_number}: titulo com mais de {MAX_TITLE_CHARS} caracteres.")
        if len(description_text) > MAX_GAME_DESCRIPTION_CHARS:
            errors.append(f"Linha {row_number}: descricao com mais de {MAX_GAME_DESCRIPTION_CHARS} caracteres.")
        truth_raw = str(row[idx["veracidade"]] or "").strip().lower()
        cards.append(
            {
                "title": title_text,
                "description": description_text,
                "score": int(row[idx["pontuacao"]] or 0),
                "truth": truth_raw in {"verdadeira", "verdadeiro", "true", "sim", "1"},
            }
        )
    if errors:
        raise ValueError("Corrija o modelo Excel antes de importar:\n" + "\n".join(errors[:10]))
    return cards
